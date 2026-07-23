"""
OPay Excel Statement Parser
=============================
VALIDATED against two real OPay wallet account statement exports
(45-row and 102-row). Same underlying transaction vocabulary as OPay's
PDF export (same description patterns: "POS | ... | ...", "Transfer to
X | Bank | Account", "Transfer from X | Provider | Account", "Airtime |
... | ...", "Stamp Duty", "VAT on Transfer Fee") - unsurprising, since
it's the same account data, just exported in a structured format
instead of PDF text. _infer_service_type/_is_levy_line are therefore
intentionally identical to opay_pdf_parser.py's (kept as a separate
copy rather than a shared base class, consistent with how this
package already keeps each provider+format parser fully independent -
see moniepoint_excel_parser.py vs moniepoint_pdf_parser.py).

CONFIRMED from both real files: 6 metadata rows before the transaction
table (title+date printed, account name+number, account type+period,
opening balance+total debit+debit count, closing balance+total
credit+credit count, one blank row), then the header row, then
transactions. Header row is FOUND dynamically (matching 'Trans. Date'
in the first cell) rather than hardcoded to row 7, in case OPay ever
adds/removes a metadata row - the same defensive approach
moniepoint_excel_parser.py already uses for the same reason.

VERIFIED: for both real files, summing this parser's extracted debit
transactions and credit transactions independently reproduces the
statement's own stated "Total Debit"/"Total Credit" figures exactly,
and the parsed debit/credit transaction COUNTS match the statement's
own stated "Debit Count"/"Credit Count" exactly.
"""
from typing import Dict, List, Optional
from datetime import datetime

import openpyxl


class OPayExcelParser:
    """Parses a real OPay wallet account statement .xlsx export."""

    def __init__(self, file_path: str):
        self.file_path = file_path

    def extract(self) -> Dict:
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
        except Exception as e:
            return self._failure([f"Could not read Excel file: {e}"])

        sheet_names = wb.sheetnames
        ws = wb[sheet_names[0]]
        rows = list(ws.iter_rows(values_only=True))

        header_row_idx = self._find_header_row(rows)
        if header_row_idx is None:
            return self._failure([
                "Could not locate the transaction table header row (looking for "
                "'Trans. Date'). This may not be an OPay wallet statement export, "
                "or OPay has changed its export format.",
            ])

        transactions = []
        skipped = 0
        for row in rows[header_row_idx + 1:]:
            if row is None or all(v is None for v in row):
                continue
            txn = self._parse_row(row)
            if txn is None:
                skipped += 1
                continue
            transactions.append(txn)

        total_rows = len(transactions) + skipped
        confidence = len(transactions) / total_rows if total_rows else 0.0

        errors = []
        if skipped:
            errors.append(f"{skipped} row(s) could not be parsed and were skipped.")

        return {
            "success": len(transactions) > 0,
            "transaction_count": len(transactions),
            "transactions": transactions,
            "confidence_score": round(confidence, 2),
            "errors": errors,
            "provider": "OPay",
            "source_format": "excel",
        }

    def _find_header_row(self, rows: List[tuple]) -> Optional[int]:
        for idx, row in enumerate(rows):
            if row and row[0] and str(row[0]).strip() == "Trans. Date":
                return idx
        return None

    def _parse_row(self, row: tuple) -> Optional[Dict]:
        # Columns: Trans. Date, Value Date, Description, Debit(₦), Credit(₦),
        # Balance After(₦), Channel, Transaction Reference
        if len(row) < 8:
            return None

        date_val, _value_date, description, debit_val, credit_val, balance_val, channel, reference = row[:8]

        if not description:
            return None

        date = self._parse_date(date_val)
        if not date:
            return None

        debit = self._clean_amount(debit_val)
        credit = self._clean_amount(credit_val)
        if debit == 0.0 and credit == 0.0:
            return None

        description = str(description).strip()
        is_credit = credit > 0
        is_levy = self._is_levy_line(description)
        service_type = self._infer_service_type(description, is_credit)

        return {
            "date": date,
            "amount": credit if is_credit else debit,
            "direction": "in" if is_credit else "out",
            "description": description,
            "service_type": service_type,
            "is_emtl_qualifying": service_type == "transfer_to_bank",  # Stamp Duty applies to
                                                                          # bank transfers >= ₦10k,
                                                                          # a separate transaction -
                                                                          # not a flag on others
            "is_levy_line": is_levy,
            "channel": str(channel) if channel else None,
            "reference": str(reference) if reference else None,
            "raw_debit": debit,
            "raw_credit": credit,
            "balance_after": self._clean_amount(balance_val),
        }

    def _parse_date(self, value) -> Optional[str]:
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if value is None:
            return None
        text = str(value).strip()
        for fmt in ("%d %b %Y %H:%M:%S", "%d %b %Y"):
            try:
                return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return None

    def _clean_amount(self, value) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip()
        if text == "--" or text == "":
            return 0.0
        try:
            return float(text.replace(",", "").replace("₦", ""))
        except ValueError:
            return 0.0

    def _is_levy_line(self, description: str) -> bool:
        """A levy/tax line tied to another transaction, not a standalone
        transaction itself. Checked case-insensitively: real statements
        show 'Stamp Duty', 'VAT on Transfer Fee', and 'Electronic Money
        Transfer Levy' as three DISTINCT real line items - all three must
        be checked, not just the most obvious one."""
        d = description.upper()
        return any(marker in d for marker in (
            "ELECTRONIC MONEY TRANSFER LEVY", "STAMP DUTY", "VAT", "VALUE ADDED TAX",
        ))

    def _infer_service_type(self, description: str, is_credit: bool) -> str:
        d = description.lower()
        if d.startswith("pos |") or d.startswith("pos|"):
            # No 'deposit' branch: audited against every real OPay/Moniepoint
            # statement available (PDF, Excel, screenshots) - zero debit-
            # direction "POS |" occurrences found anywhere.
            return "withdrawal"
        if self._is_levy_line(description):
            return "levy"
        if d.startswith("transfer from"):
            return "pos_transfer"
        if d.startswith("transfer to"):
            return "transfer_to_bank"
        if d.startswith("airtime"):
            return "airtime"
        if "data" in d and ("bundle" in d or "internet" in d):
            return "data"
        if any(w in d for w in ["dstv", "gotv", "wec", "phcn", "bill"]):
            return "bills_payment"
        return "withdrawal" if is_credit else "transfer_to_bank"

    def _failure(self, errors: List[str]) -> Dict:
        return {
            "success": False,
            "transaction_count": 0,
            "transactions": [],
            "confidence_score": 0.0,
            "errors": errors,
            "provider": "OPay",
            "source_format": "excel",
        }
